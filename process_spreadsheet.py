import openpyxl as xl
wb=xl.load_workbook('transactions.xlsx')
sheet=wb['Sheet1']

for row in range(2,sheet.max_row+1):
    sheet.cell(row,4).value=sheet.cell(row,3).value*0.9
wb.save('transactions2.xlsx')